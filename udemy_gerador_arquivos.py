import streamlit as st
import pandas as pd
import io
import re
import csv
from datetime import datetime
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path as PathlibPath
from openai import OpenAI
import time
import json

CSV_HEADER = [
    "Question", "Question Type",
    "Answer Option 1", "Explanation 1",
    "Answer Option 2", "Explanation 2",
    "Answer Option 3", "Explanation 3",
    "Answer Option 4", "Explanation 4",
    "Answer Option 5", "Explanation 5",
    "Answer Option 6", "Explanation 6",
    "Correct Answers", "Overall Explanation", "Domain"
]

ASSISTANT_ID = "asst_5TeFXS410FdC2LZvAvOIqa96"
client = OpenAI()

def aguardar_resposta(thread_id, run_id, timeout=60):
    for _ in range(timeout):
        run_status = client.beta.threads.runs.retrieve(thread_id=thread_id, run_id=run_id)
        if run_status.status == "completed":
            return True
        elif run_status.status in ["failed", "cancelled"]:
            return False
        time.sleep(1)
    return False

def processar_questoes(texto, origem):
    questoes = []
    blocos = re.split(r'Question \d+', texto)

    total = len([b for b in blocos if b.strip()])
    atual = 1

    for bloco in blocos:
        bloco = bloco.strip()
        if not bloco:
            continue

        with st.spinner(f"🔄 Processando questão {atual} de {total} com IA..."):
            question_data = gerar_pergunta_xlsx_com_ia(bloco)
            st.code(question_data)

        if question_data:
            question_data["Origem"] = origem
            questoes.append(question_data)

        atual += 1

    return questoes

def gerar_pergunta_xlsx_com_ia(bloco):
    prompt = f"""
Extraia a seguinte estrutura JSON da questão abaixo. Se houver listas no enunciado (por exemplo, com marcadores como •, -, * ou números), preserve os itens em nova linha para facilitar a leitura.

Retorne no seguinte formato:
{{
  "Pergunta": "...",
  "Opção A": "...",
  "Opção B": "...",
  "Opção C": "...",
  "Opção D": "...",
  "Opção E": "...",
  "Resposta(s) Correta(s)": "Texto(s) exato(s) da(s) resposta(s) correta(s)",
  "Explicação": "..."
}}

Texto:
{bloco.strip()}
"""
    try:
        thread = client.beta.threads.create()
        thread_id = thread.id

        client.beta.threads.messages.create(
            thread_id=thread_id,
            role="user",
            content=prompt
        )

        run = client.beta.threads.runs.create(thread_id=thread_id, assistant_id=ASSISTANT_ID)
        if not aguardar_resposta(thread_id, run.id):
            st.error("❌ A execução da IA falhou ou foi cancelada.")
            return None

        final_msg = client.beta.threads.messages.list(thread_id=thread_id).data[0].content[0].text.value
        st.code(final_msg)  # Para debug visual
        try:
            return json.loads(final_msg)
        except json.JSONDecodeError:
            return None

    except Exception as e:
        st.error(f"Erro ao processar com IA: {e}")
        return None

def gerar_xlsx(questoes, nome_arquivo):
    df = pd.DataFrame(questoes)
    caminho = f"{nome_arquivo}.xlsx"
    with pd.ExcelWriter(caminho, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="Questoes", index=False)
        wb = writer.book
        ws = writer.sheets["Questoes"]
        for col_num, _ in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 30

def gerar_csv_udemy(texto):
    output = io.StringIO()
    questions = []
    blocks = re.split(r"(?=Question \d+\n)", texto.strip())

    total = len(blocks)
    for idx, block in enumerate(blocks, 1):
        with st.spinner(f"🔄 Processando questão {idx} de {total} com IA..."):
            question_data = gerar_pergunta_csv_com_ia(block)
            st.code(question_data)

        if question_data:
            questions.append(question_data)

    df_csv = pd.DataFrame(questions, columns=CSV_HEADER)
    df_csv.to_csv(output, index=False)
    output.seek(0)
    return output.getvalue(), len(df_csv)

def gerar_pergunta_csv_com_ia(bloco):
    prompt = f"""
Você é um assistente especialista em gerar arquivos CSV para a Udemy com base em perguntas de múltipla escolha.

Extraia da pergunta abaixo a estrutura JSON exatamente no seguinte formato (não adicione comentários nem quebras extras):

{{
  "Question": "...",
  "Question Type": "multiple-choice" ou "multi-select",
  "Answer Option 1": "...",
  "Answer Option 2": "...",
  "Answer Option 3": "...",
  "Answer Option 4": "...",
  "Answer Option 5": "...",
  "Answer Option 6": "...",
  "Correct Answers": "1" ou "1,3",
  "Overall Explanation": "...",
  "Domain": ""
}}

Regras:
- Separe o enunciado da pergunta e as opções corretamente.
- Sempre preencha no mínimo 2 opções de resposta.
- Se houver uma ou mais respostas corretas, use números (1 a 6) no campo "Correct Answers".
- Nunca coloque as opções dentro do texto da pergunta.
- Retorne apenas o JSON e nada mais.

Pergunta:
{bloco.strip()}
"""
    try:
        thread = client.beta.threads.create()
        thread_id = thread.id

        client.beta.threads.messages.create(
            thread_id=thread_id,
            role="user",
            content=prompt
        )

        run = client.beta.threads.runs.create(thread_id=thread_id, assistant_id=ASSISTANT_ID)
        if not aguardar_resposta(thread_id, run.id):
            st.error("❌ A execução da IA falhou ou foi cancelada.")
            return None

        final_msg = client.beta.threads.messages.list(thread_id=thread_id).data[0].content[0].text.value
        st.markdown("🧠 **Resposta recebida da IA:**")
        st.code(final_msg, language="json")
        try:
            return json.loads(final_msg)
        except json.JSONDecodeError:
            return None

    except Exception as e:
        st.error(f"Erro ao processar com IA: {e}")
        return None

def agregar_planilhas(uploaded_files):
    frames = []
    for file in uploaded_files:
        try:
            df = pd.read_excel(file)
            frames.append(df)
        except Exception as e:
            st.error(f"Erro ao ler o arquivo {file.name}: {e}")

    if not frames:
        st.warning("Nenhuma planilha válida foi carregada.")
        return

    df_final = pd.concat(frames, ignore_index=True)
    buffer = io.BytesIO()
    df_final.to_excel(buffer, index=False)
    buffer.seek(0)

    st.success(f"✅ {len(df_final)} questões agregadas com sucesso!")
    st.download_button(
        label="📥 Baixar Planilha Agregada",
        data=buffer,
        file_name="planilha_agregada.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
